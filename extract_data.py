import unittest
import openpyxl
import datetime

class excel_data(unittest.TestCase):
    wb = openpyxl.load_workbook("CPD.xlsx")
    sheet = wb["Manual"]

    def data_extract(self):
        self.data_check()
        list_string = []
        for k in range(4,len(self.sheet['B']) +1):
            for i in range(2,11,2):
                if i is not None:
                    list_string += [self.sheet.cell(row=k, column=i).value]
        list_string = [str(x)[:10] if type(x) == datetime.datetime else x for x in list_string] #replace datetime type with string (times are stripped)
        return list_string

    def data_complete(self):
        list_string = self.data_extract()
        for i in range(4, len(list_string),5):
            if list_string[i] == None:
                list_string[i] = ""
        return list_string

    def data_check(self):
        #check if data is appropriate
        for j in range(4, len(self.sheet['B'])):
            for t in range(2,8,2):
                # no blanks
                if self.sheet.cell(row=j, column=t).value == None:
                    print("You shouldn't leave any cells blank for Verifiability, Number of hours,\
                    Activity Name, and Date earned")
                    raise SystemExit
                #only interger acceptable for hours
                elif not str(self.sheet.cell(row=j, column=4).value).isnumeric():
                    print('There (is)are non-number values in the Number of Hourse column')
                    raise SystemExit
                #only date format acceptable
                elif type(self.sheet.cell(row=j, column=8).value) != datetime.datetime:
                    print('There (is)are non-date values in the Date Earned column')
                    raise SystemExit

#
# if __name__ == '__main__':
#     unittest.main()

# list_string = []
# for i in range(4,13,2):
#     if i is not None:
#         print(type(i))
# for k in range(4, len(sheet['B'])):
#     list_string += [sheet.cell(row=k, column=i).value for i in range(4,11) if sheet.cell(row=k, column=i).value != None]
# list_string = [str(x)[:10] if type(x) == datetime.datetime else x for x in list_string]
# print(list_string)
#
# print(len(sheet['H']))
#
# sheet['B7'].value = ''
# sheet['D7'].value = 1

# print(sheet['D7'].value )
#
# wb.save("CPD.xlsx")
